"""XLSX -> ParsedDocument.

Uses openpyxl. Each worksheet becomes a HeadingBlock (sheet name) followed by one
TableBlock covering the sheet's used range, followed by any pictures embedded on
that sheet. Merged cell ranges map directly to `TableData.merges`; openpyxl
already reports covered cells as None, matching the IR convention.

Byte offsets (Karar B, best-effort): the TableBlock's Span points at the
`<sheetData>` element inside `xl/worksheets/sheetN.xml` when that part can be
located and parsed; the ImageBlock's Span likewise points at that sheet's
`<drawing>` reference element; otherwise the byte offsets are left unset.

Embedded pictures are read directly from the zip (drawingML relationships:
sheetN.xml -> its drawing part -> the drawing's own media relationships),
NOT via openpyxl's `Worksheet._images` -- that path silently returns nothing
when Pillow isn't installed (openpyxl's `find_images` early-returns if
`PILImage` failed to import) and doesn't retain the media part path anyway
(`Image.ref` is a BytesIO, not "xl/media/imageN.png"). Native charts
(c:chart) and SmartArt-equivalents are out of scope here too, same call as
pptx's chart/SmartArt gap (see EKSIKLER) -- a `graphicFrame` anchor with no
`xdr:pic` child is simply skipped.
"""

from __future__ import annotations

import hashlib
import posixpath
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl

from .base import (
    BaseParser, ParsedDocument, Span, HeadingBlock, TableBlock, TableData, Merge,
    ImageBlock, text_cell,
)
from ._ooxml import element_byte_range

_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_DRAWING_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing")
_ANCHOR_TAGS = {"oneCellAnchor", "twoCellAnchor", "absoluteAnchor"}
_MIME_BY_EXT = {
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".gif": "image/gif", ".bmp": "image/bmp", ".tif": "image/tiff",
    ".tiff": "image/tiff", ".emf": "image/x-emf", ".wmf": "image/x-wmf",
}
_SHEET_NUM = re.compile(r"sheet(\d+)\.xml$")


def _cell(value) -> str | None:
    if value is None:
        return None
    return str(value)


def _parse_rels(zf: zipfile.ZipFile, part: str) -> dict[str, tuple[str, str]]:
    """`part`'s _rels sidecar as {id: (type, resolved_target_part)}; {} if none."""
    rels_path = posixpath.join(posixpath.dirname(part), "_rels",
                                posixpath.basename(part) + ".rels")
    if rels_path not in zf.namelist():
        return {}
    tree = ET.fromstring(zf.read(rels_path))
    base = posixpath.dirname(part)
    out: dict[str, tuple[str, str]] = {}
    for rel in tree.findall(f"{{{_NS_PKG_REL}}}Relationship"):
        if rel.get("TargetMode") == "External":
            continue
        target = rel.get("Target", "")
        # Some writers (e.g. openpyxl) emit a zip-root-absolute target
        # ("/xl/drawings/drawing1.xml") instead of a relative one; only
        # relative targets resolve against `base`.
        resolved = (posixpath.normpath(target.lstrip("/")) if target.startswith("/")
                    else posixpath.normpath(posixpath.join(base, target)))
        out[rel.get("Id")] = (rel.get("Type", ""), resolved)
    return out


def _sheet_images(zf: zipfile.ZipFile, sheet_part: str) -> list[dict]:
    """Pictures anchored on one worksheet: [{"part", "mime", "row", "col"}, ...].

    Charts and any other non-`xdr:pic` anchor content are skipped (see module
    docstring) -- only real embedded raster/vector pictures are extracted.
    """
    sheet_rels = _parse_rels(zf, sheet_part)
    drawing_part = next((tgt for (typ, tgt) in sheet_rels.values()
                         if typ == _DRAWING_REL_TYPE), None)
    if not drawing_part or drawing_part not in zf.namelist():
        return []
    drawing_rels = _parse_rels(zf, drawing_part)
    try:
        dtree = ET.fromstring(zf.read(drawing_part))
    except ET.ParseError:
        return []

    out: list[dict] = []
    for anchor in dtree:
        if anchor.tag.rsplit("}", 1)[-1] not in _ANCHOR_TAGS:
            continue
        pic = anchor.find(f"{{{_NS_XDR}}}pic")
        if pic is None:
            continue
        blip = pic.find(f".//{{{_NS_A}}}blip")
        embed = blip.get(f"{{{_NS_R}}}embed") if blip is not None else None
        if not embed or embed not in drawing_rels:
            continue
        _typ, media_part = drawing_rels[embed]
        row = col = None
        frm = anchor.find(f"{{{_NS_XDR}}}from")
        if frm is not None:
            row_el = frm.find(f"{{{_NS_XDR}}}row")
            col_el = frm.find(f"{{{_NS_XDR}}}col")
            if row_el is not None and row_el.text:
                row = int(row_el.text)
            if col_el is not None and col_el.text:
                col = int(col_el.text)
        ext = posixpath.splitext(media_part)[1].lower()
        out.append({"part": media_part, "mime": _MIME_BY_EXT.get(ext),
                    "row": row, "col": col})
    return out


class XlsxParser(BaseParser):
    extensions = (".xlsx",)
    mimetypes = (_MIME,)
    fmt = "xlsx"
    version = "xlsx/0.1"

    def parse(self, raw_path: str | Path, doc_id: str) -> ParsedDocument:
        raw_path = Path(raw_path)
        data = raw_path.read_bytes()
        raw_sha256 = hashlib.sha256(data).hexdigest()

        wb = openpyxl.load_workbook(raw_path, data_only=True)
        # Raw sheet XML parts, in workbook order: byte-range + embedded pictures.
        sheet_meta = self._sheet_meta(raw_path)

        blocks: list = []
        bid = 0
        img_n = 0

        def next_id() -> str:
            nonlocal bid
            s = f"b{bid}"
            bid += 1
            return s

        for idx, ws in enumerate(wb.worksheets):
            meta = sheet_meta.get(idx, {})
            images = meta.get("images", [])

            rows: list[list[str | None]] = []
            if ws.max_row is not None and ws.max_column is not None:
                min_r, min_c = ws.min_row, ws.min_column
                max_r, max_c = ws.max_row, ws.max_column
                rows = [
                    [_cell(v) for v in row]
                    for row in ws.iter_rows(min_row=min_r, max_row=max_r,
                                            min_col=min_c, max_col=max_c,
                                            values_only=True)
                ]
            has_data = bool(rows) and not all(c is None for r in rows for c in r)
            if not has_data and not images:
                continue  # nothing on this sheet

            blocks.append(HeadingBlock(id=next_id(), text=ws.title, level=1))
            part = meta.get("part")

            if has_data:
                merges = []
                for rng in ws.merged_cells.ranges:
                    merges.append(Merge(
                        row=rng.min_row - min_r, col=rng.min_col - min_c,
                        rowspan=rng.max_row - rng.min_row + 1,
                        colspan=rng.max_col - rng.min_col + 1))
                table = TableData(
                    n_rows=len(rows), n_cols=max_c - min_c + 1,
                    cells=[[text_cell(v) for v in row] for row in rows],
                    merges=merges)
                brange = meta.get("table_range")
                span = Span(part=part,
                            byte_start=brange[0] if brange else None,
                            byte_end=brange[1] if brange else None)
                blocks.append(TableBlock(id=next_id(), span=span, table=table))

            if images:
                drange = meta.get("drawing_range")
                dspan = Span(part=part,
                             byte_start=drange[0] if drange else None,
                             byte_end=drange[1] if drange else None)
                for img in images:
                    img_n += 1
                    locator = {"part": img["part"]}
                    if img["row"] is not None:
                        locator["anchor_row"] = img["row"]
                    if img["col"] is not None:
                        locator["anchor_col"] = img["col"]
                    blocks.append(ImageBlock(
                        id=next_id(), span=dspan, image_index=img_n,
                        locator=locator, mime=img["mime"]))

        return ParsedDocument(
            doc_id=doc_id,
            source_path=str(raw_path),
            fmt=self.fmt,
            raw_sha256=raw_sha256,
            mimetype=_MIME,
            parser_version=self.version,
            metadata={"sheets": [ws.title for ws in wb.worksheets]},
            blocks=blocks,
        )

    @staticmethod
    def _sheet_meta(raw_path: Path) -> dict[int, dict]:
        """Best-effort, positional (index -> xl/worksheets/sheetN.xml) sheet
        metadata: {"part", "table_range", "drawing_range", "images"}.

        Sheets are matched positionally in NUMERIC sheetN.xml order (not
        lexicographic -- "sheet10.xml" must sort after "sheet2.xml", not before
        "sheet2.xml"). This is a heuristic (reordered/deleted sheets can shift
        it), so failures just yield no offsets/images rather than wrong ones.
        """
        out: dict[int, dict] = {}
        try:
            with zipfile.ZipFile(raw_path) as z:
                names = sorted(
                    (n for n in z.namelist()
                     if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")),
                    key=lambda n: int(m.group(1)) if (m := _SHEET_NUM.search(n)) else 0)
                for idx, name in enumerate(names):
                    xml = z.read(name)
                    out[idx] = {
                        "part": name,
                        "table_range": element_byte_range(xml, "sheetData"),
                        "drawing_range": element_byte_range(xml, "drawing"),
                        "images": _sheet_images(z, name),
                    }
        except (zipfile.BadZipFile, KeyError, OSError):
            pass
        return out
