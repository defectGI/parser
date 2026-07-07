"""xlsx_parser.py: embedded pictures (xl/drawings) extraction and the
numeric (not lexicographic) sheetN.xml ordering fix.
"""

from __future__ import annotations

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from parsers.base import HeadingBlock, ImageBlock, TableBlock
from parsers.xlsx_parser import XlsxParser


def _png(tmp_path, name="pic.png"):
    p = tmp_path / name
    PILImage.new("RGB", (4, 4), color="red").save(p)
    return str(p)


def test_basic_sheet_and_merges(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Header"
    ws["A2"] = "Value"
    ws.merge_cells("A1:B1")
    xlsx = tmp_path / "basic.xlsx"
    wb.save(xlsx)

    doc = XlsxParser().parse(xlsx, "doc")
    heading = next(b for b in doc.blocks if isinstance(b, HeadingBlock))
    table = next(b for b in doc.blocks if isinstance(b, TableBlock))
    assert heading.text == "Data"
    assert table.table.cells[0][0].plain_text() == "Header"
    assert len(table.table.merges) == 1


def test_embedded_image_extracted(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "x"
    ws.add_image(XLImage(_png(tmp_path)), "C2")
    xlsx = tmp_path / "img.xlsx"
    wb.save(xlsx)

    doc = XlsxParser().parse(xlsx, "doc")
    images = [b for b in doc.blocks if isinstance(b, ImageBlock)]
    assert len(images) == 1
    assert images[0].locator["part"] == "xl/media/image1.png"
    assert images[0].locator["anchor_row"] == 1   # C2 -> 0-based row 1
    assert images[0].locator["anchor_col"] == 2   # C2 -> 0-based col 2
    assert images[0].mime == "image/png"


def test_image_only_sheet_not_dropped(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "x"
    pic_only = wb.create_sheet("PicOnly")
    pic_only.add_image(XLImage(_png(tmp_path)), "A1")
    xlsx = tmp_path / "piconly.xlsx"
    wb.save(xlsx)

    doc = XlsxParser().parse(xlsx, "doc")
    headings = [b.text for b in doc.blocks if isinstance(b, HeadingBlock)]
    assert "PicOnly" in headings
    pic_idx = headings.index("PicOnly")
    # the block right after the PicOnly heading is its image, not a TableBlock
    idx = doc.blocks.index(next(b for b in doc.blocks
                                if isinstance(b, HeadingBlock) and b.text == "PicOnly"))
    assert isinstance(doc.blocks[idx + 1], ImageBlock)


def test_many_sheets_numeric_order(tmp_path):
    # 11 sheets so "sheet10.xml"/"sheet11.xml" would sort before "sheet2.xml"
    # under a naive lexicographic string sort.
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    wb.active["A1"] = "v1"
    for i in range(2, 12):
        ws = wb.create_sheet(f"S{i}")
        ws["A1"] = f"v{i}"
    xlsx = tmp_path / "many.xlsx"
    wb.save(xlsx)

    doc = XlsxParser().parse(xlsx, "doc")
    tables = [b for b in doc.blocks if isinstance(b, TableBlock)]
    assert len(tables) == 11
    for i, t in enumerate(tables, start=1):
        assert t.table.cells[0][0].plain_text() == f"v{i}"
        assert t.span.part == f"xl/worksheets/sheet{i}.xml"
